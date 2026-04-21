import streamlit as st
import numpy as np
import re
import io
import base64
import json
import requests
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── GPT-4o-mini Vision ────────────────────────────────────────────────────────

def imagen_a_base64(img_pil, quality=90):
    buf = io.BytesIO()
    img_pil.save(buf, format="JPEG", quality=quality)
    return base64.b64encode(buf.getvalue()).decode()

def recortar_marca_agua(img_pil):
    """Recorta SOLO la esquina inferior derecha donde está la marca de agua azul.
    Elimina completamente la pantalla del GPS del contexto visual."""
    w, h = img_pil.size
    return img_pil.crop((int(w * 0.35), int(h * 0.86), w, h))

def llamar_gpt(img_pil, api_key, modelo="gpt-4o-mini", detalle="high"):
    """Siempre recorta la zona de marca de agua antes de enviar a GPT."""
    img_recortada = recortar_marca_agua(img_pil)

    payload = {
        "model": modelo,
        "max_tokens": 150,
        "temperature": 0,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{imagen_a_base64(img_recortada)}",
                        "detail": detalle
                    }
                },
                {
                    "type": "text",
                    "text": (
                        "Esta imagen muestra solo la marca de agua azul de una foto GPS Garmin.\n"
                        "Contiene dos líneas de texto azul con este formato:\n"
                        "  DD mes AAAA  H:MM:SS a.m./p.m.\n"
                        "  XX.XXXXXXS  YY.YYYYYYYW\n\n"
                        "Ejemplo:\n"
                        "  8 abr 2026  10:26:27 a.m.\n"
                        "  14.005450S  69.252968W\n\n"
                        "Lee el texto EXACTAMENTE como aparece y responde SOLO con JSON (sin markdown):\n"
                        '{"fecha":"8 abr 2026","hora":"10:26:27 a.m.",'
                        '"latitud":"14.005450S","longitud":"69.252968W"}\n'
                        "Si no puedes leer algún valor usa null."
                    )
                }
            ]
        }]
    }

    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {api_key}"},
        json=payload, timeout=30
    )

    if resp.status_code != 200:
        return None

    texto = resp.json()["choices"][0]["message"]["content"].strip()
    texto = re.sub(r"```json|```", "", texto).strip()
    return json.loads(texto)


def parsear_lat(s):
    if not s: return None
    m = re.search(r'([\d.]+)\s*([NS])', str(s).upper())
    if m:
        v = float(m.group(1))
        return -v if m.group(2) == 'S' else v
    return None

def parsear_lon(s):
    if not s: return None
    m = re.search(r'([\d.]+)\s*([WEO])', str(s).upper())
    if m:
        v = float(m.group(1))
        return -v if m.group(2) in ('W','O') else v
    return None

def convertir_utm(data):
    """Convierte coordenadas UTM a decimales."""
    try:
        import utm as utm_lib
        for val in data.values():
            if val:
                m = re.search(r'(\d{1,2})([A-Z])\s+(\d{6,7})\s+(\d{7})', str(val).upper())
                if m:
                    lat, lon = utm_lib.to_latlon(
                        int(m.group(3)), int(m.group(4)),
                        int(m.group(1)), m.group(2)
                    )
                    return round(lat, 6), round(lon, 6)
    except Exception:
        pass
    return None, None

def llamar_gpt_multiformat(img_pil, api_key, modelo="gpt-4o-mini"):
    """Detecta coordenadas decimales y UTM."""
    img_recortada = recortar_marca_agua(img_pil)
    payload = {
        "model": modelo,
        "max_tokens": 200,
        "temperature": 0,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{imagen_a_base64(img_recortada)}",
                        "detail": "high"
                    }
                },
                {
                    "type": "text",
                    "text": (
                        "Lee el texto de la marca de agua GPS visible en esta imagen.\n"
                        "Responde SOLO con JSON (sin markdown).\n\n"
                        "Si las coordenadas son DECIMALES (ej: 14.062602S 69.203552W):\n"
                        '{"tipo":"decimal","fecha":"8 abr 2026","hora":"10:26:27 a.m.",'
                        '"latitud":"14.005450S","longitud":"69.252968W","utm":null}\n\n'
                        "Si las coordenadas son UTM (ej: 19L 248384 8454372):\n"
                        '{"tipo":"utm","fecha":"18 abr 2026","hora":"8:47:30 a. m.",'
                        '"latitud":null,"longitud":null,"utm":"19L 248384 8454372"}\n\n'
                        "Extrae los valores EXACTOS. Si no puedes leer alguno usa null."
                    )
                }
            ]
        }]
    }
    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {api_key}"},
        json=payload, timeout=30
    )
    if resp.status_code != 200:
        return None
    texto = resp.json()["choices"][0]["message"]["content"].strip()
    texto = re.sub(r"```json|```", "", texto).strip()
    return json.loads(texto)


def extraer_coordenadas(img_pil, api_key):
    """gpt-4o-mini primero, reintenta con gpt-4o si falla."""
    for modelo in ["gpt-4o-mini", "gpt-4o"]:
        data = llamar_gpt_multiformat(img_pil, api_key, modelo=modelo)
        if not data:
            continue
        fecha = data.get("fecha")
        hora  = data.get("hora")
        if data.get("tipo") == "utm":
            lat, lon = convertir_utm(data)
        else:
            lat = parsear_lat(data.get("latitud"))
            lon = parsear_lon(data.get("longitud"))
            if lat is None or lon is None:
                lat2, lon2 = convertir_utm(data)
                if lat is None: lat = lat2
                if lon is None: lon = lon2
        if lat is not None and lon is not None:
            return lat, lon, fecha, hora
    return None, None, None, None


# ── Excel ─────────────────────────────────────────────────────────────────────

def generar_excel(datos):
    wb = Workbook(); ws = wb.active; ws.title = "Coordenadas GPS"
    ft = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    fn = Font(name='Arial', size=10)
    fa = PatternFill('solid', start_color='2E75B6')
    fb = PatternFill('solid', start_color='DDEEFF')
    cc = Alignment(horizontal='center', vertical='center')
    ci = Alignment(horizontal='left',   vertical='center')
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'))

    for col, enc in enumerate(['#','Archivo','Fecha','Hora','Latitud','Longitud','Estado'], 1):
        c = ws.cell(row=1, column=col, value=enc)
        c.font=ft; c.fill=fa; c.alignment=cc; c.border=bd
    ws.row_dimensions[1].height = 22

    for i, (arch, fecha, hora, lat, lon, est) in enumerate(datos, 1):
        # Convertir lat/lon a texto con punto decimal para evitar problema regional
        lat_str = f"{lat:.6f}".replace(',', '.') if lat is not None else ""
        lon_str = f"{lon:.6f}".replace(',', '.') if lon is not None else ""
        fila_vals = [i, arch, fecha, hora, lat_str, lon_str, est]
        for col, val in enumerate(fila_vals, 1):
            c = ws.cell(row=i+1, column=col, value=val)
            c.font=fn; c.alignment=ci if col==2 else cc; c.border=bd
            if i%2==0: c.fill=fb

    for col, ancho in zip('ABCDEFG', [5,38,18,16,14,14,16]):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Extractor GPS Garmin", page_icon="📍", layout="wide")

st.markdown("""
<style>
.main-title{font-size:2.2rem;font-weight:700;color:#1a3a5c;}
.subtitle{color:#666;font-size:1rem;margin-bottom:1rem;}
.costo{background:#f0fff4;border:1px solid #68d391;border-radius:8px;
       padding:0.5rem 1rem;font-size:0.9rem;color:#276749;margin-bottom:1rem;}
.reintento{background:#fffbeb;border:1px solid #f6ad55;border-radius:6px;
           padding:0.3rem 0.8rem;font-size:0.8rem;color:#744210;}
</style>""", unsafe_allow_html=True)

st.markdown('<p class="main-title">📍 Extractor de Coordenadas GPS</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Sube tus fotos Garmin · Extrae coordenadas automáticamente · Descarga en Excel</p>', unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🔑 API Key de OpenAI")
    api_key = st.text_input(
        "Pega tu API Key aquí",
        type="password",
        placeholder="sk-...",
        help="Obtén tu key en: platform.openai.com/api-keys"
    )
    if api_key:
        st.success("✅ API Key ingresada")
    else:
        st.warning("⚠️ Ingresa tu API Key")
        st.markdown("[Obtener API Key →](https://platform.openai.com/api-keys)")

    st.divider()
    st.markdown("### 💰 Costo estimado")
    st.markdown("""
    | Fotos | Costo aprox. |
    |-------|-------------|
    | 10    | ~$0.01      |
    | 50    | ~$0.05      |
    | 100   | ~$0.10      |
    | 500   | ~$0.50      |

    Usa **gpt-4o-mini** por defecto.
    Si no detecta lat/lon, reintenta con **gpt-4o** automáticamente.
    """)
    st.divider()
    st.markdown("### ℹ️ Cómo usar")
    st.markdown("1. 🔑 Ingresa tu API Key\n2. 📤 Sube las fotos\n3. 🔍 Clic en **Extraer**\n4. 📥 Descarga el **Excel**")
    st.divider()
    st.caption("v2.1 · Extractor GPS Garmin")

# ── Main ──────────────────────────────────────────────────────────────────────
if not api_key:
    st.info("👈 Ingresa tu API Key de OpenAI en el panel izquierdo para comenzar.")
    st.stop()

archivos = st.file_uploader(
    "📤 Arrastra tus fotos o haz clic para seleccionar",
    type=["jpg","jpeg","png"],
    accept_multiple_files=True
)

if not archivos:
    st.info("👆 Sube tus fotos para comenzar.")
    st.stop()

costo_est = len(archivos) * 0.0012
st.markdown(
    f'<div class="costo">💰 Costo estimado para <b>{len(archivos)}</b> foto(s): '
    f'<b>~${costo_est:.3f} USD</b> (puede subir levemente si hay reintentos)</div>',
    unsafe_allow_html=True
)

with st.expander(f"🖼️ Vista previa — {len(archivos)} foto(s)", expanded=False):
    cols = st.columns(min(len(archivos), 4))
    for i, f in enumerate(archivos):
        with cols[i%4]:
            f.seek(0)
            st.image(Image.open(f), caption=f.name, use_container_width=True)

st.divider()

if st.button("🔍 Extraer coordenadas", type="primary", use_container_width=True):
    datos=[]; ok_count=0; reintentos=0
    progress = st.progress(0, text="Iniciando...")
    status   = st.empty()

    hcols = st.columns([0.4, 2.2, 1.4, 1.4, 1.4, 1.4, 0.5])
    for col, lbl in zip(hcols, ['#','Archivo','Fecha','Hora','Latitud','Longitud','✓']):
        col.markdown(f"**{lbl}**")
    st.markdown("---")

    for idx, archivo in enumerate(archivos):
        status.info(f"⏳ Procesando **{archivo.name}** ({idx+1}/{len(archivos)})")
        try:
            archivo.seek(0)
            img = Image.open(archivo).convert("RGB")
            lat, lon, fecha, hora = extraer_coordenadas(img, api_key)

            icono = "✅" if (lat and lon) else "⚠️"
            if lat and lon: ok_count += 1

            row = st.columns([0.4, 2.2, 1.4, 1.4, 1.4, 1.4, 0.5])
            row[0].write(idx+1)
            row[1].write(archivo.name)
            row[2].write(fecha or "—")
            row[3].write(hora  or "—")
            row[4].write(f"{lat:.6f}" if lat is not None else "—")
            row[5].write(f"{lon:.6f}" if lon is not None else "—")
            row[6].write(icono)

            datos.append((archivo.name, fecha, hora, lat, lon,
                          "OK" if lat and lon else "Sin coordenadas"))

        except Exception as e:
            row = st.columns([0.4, 2.2, 1.4, 1.4, 1.4, 1.4, 0.5])
            row[0].write(idx+1); row[1].write(archivo.name); row[6].write("❌")
            datos.append((archivo.name, None, None, None, None, str(e)))

        progress.progress((idx+1)/len(archivos),
                          text=f"Procesando {idx+1}/{len(archivos)}...")

    status.empty(); progress.empty()
    st.divider()

    if ok_count > 0:
        fail = len(archivos) - ok_count
        st.success(f"✅ {ok_count} foto(s) procesadas · {fail} sin detectar")
        st.download_button(
            label="📥 Descargar Excel con coordenadas",
            data=generar_excel(datos),
            file_name="coordenadas_gps.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True
        )
    else:
        st.error("❌ No se detectaron coordenadas. Verifica tu API Key y las fotos.")
